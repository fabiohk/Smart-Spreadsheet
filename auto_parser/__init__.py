import pandas as pd
from openpyxl import load_workbook

from .functions import parse_tables_from_df


class AutoParser:
    def parse_xlsx(self, file_path: str) -> list[pd.DataFrame]:
        wb = load_workbook(file_path, data_only=True)

        tables = []
        for sheet_name in wb.sheetnames:
            tables_from_sheet = parse_tables_from_df(
                pd.DataFrame(wb[sheet_name].values)
            )
            tables.extend(tables_from_sheet)

        return tables
