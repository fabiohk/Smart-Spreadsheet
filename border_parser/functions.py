import numpy as np
import pandas as pd
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from manual_parser.functions import pick_table
from manual_parser.types import TableInfo


def is_border_styled(cell: Cell):
    cell_border = cell.border
    return (
        cell_border.left.style
        or cell_border.top.style
        or cell_border.right.style
        or cell_border.bottom.style
    )


def is_offlimits(sheet: Worksheet, row: int, col: int) -> bool:
    return row < 1 or row > sheet.max_row or col < 1 or col > sheet.max_column


def dfs(sheet: Worksheet, cell: Cell, visited_coordinates: set[str]) -> list[str]:
    visited_neighbors = []
    stack = [cell]
    directions = [(0, -1), (-1, 0), (0, 1), (1, 0)]
    while stack:
        current_cell = stack.pop()
        if current_cell.coordinate in visited_coordinates:
            continue
        visited_coordinates.add(current_cell.coordinate)
        visited_neighbors.append(current_cell.coordinate)
        row, col = current_cell.row, current_cell.column
        for row_diff, col_diff in directions:
            new_row, new_col = row + row_diff, col + col_diff

            if is_offlimits(sheet, new_row, new_col):
                continue

            neighbor = sheet.cell(row=new_row, column=new_col)
            if (
                is_border_styled(neighbor)
                and neighbor.coordinate not in visited_coordinates
            ):
                stack.append(neighbor)

    return visited_neighbors


def parse_tables_from_sheet(sheet: Worksheet):
    visited_coordinates = set()
    num_tables = 0
    tables_info = []
    for row in sheet.iter_rows():
        for cell in row:
            if is_border_styled(cell) and cell.coordinate not in visited_coordinates:
                num_tables += 1
                tables_info.append(dfs(sheet, cell, visited_coordinates))

    tables = []
    for table in tables_info:
        rows = [sheet[coordinate].row for coordinate in table]
        min_row, max_row = min(rows), max(rows)
        cols = [sheet[coordinate].column for coordinate in table]
        min_col, max_col = min(cols), max(cols)
        tables.append(create_df(sheet, min_row, max_row, min_col, max_col))

    return tables


def create_df(
    sheet: Worksheet, start_row: int, end_row: int, start_col: int, end_col: int
) -> pd.DataFrame:
    topleft_cell = sheet.cell(start_row, start_col).coordinate
    num_rows = end_row - start_row + 1
    num_cols = end_col - start_col + 1
    return pick_table(sheet, TableInfo("", topleft_cell, num_rows, num_cols))
